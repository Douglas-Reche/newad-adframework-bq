#!/usr/bin/env python3
"""
IO Plan Drive Sync
------------------
Scans Google Drive CLIENT/ANO/MES/PLANO/ for each active client,
processes all .xlsx files found, and loads into:
  raw.io_plan_drive_snapshot  (strategy grain, all files)
  core.io_plan_manual         (flight grain, aggregated from raw)

Usage:
  python sync_drive.py                  # sync all active clients
  python sync_drive.py --client cora    # sync specific client
  python sync_drive.py --force          # re-sync even if file unchanged

Auth:
  Local  : OAuth Desktop flow (C:/Temp/io-plan-oauth-client.json -> token cached in
           C:/Temp/io-plan-token.json). First run opens browser, subsequent runs silent.
  Cloud Run: service account via workload identity (no key file needed).
  Override: set GOOGLE_APPLICATION_CREDENTIALS=path/to/sa-key.json for service-account locally.
"""

import argparse
import io
import logging
import os
import re
import sys
import calendar
from datetime import date, datetime, timezone
from typing import Optional

import openpyxl
from google.auth import default as google_auth_default
from google.cloud import bigquery
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ─── CONFIG ───────────────────────────────────────────────────────────────────

DRIVE_ROOT_FOLDER = "0ACFCcMtN5j8EUk9PVA"
BQ_PROJECT = "adframework"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Nomes de pessoas que aparecem nos arquivos pessoais/rascunho.
# Quando uma pasta PLANO tem múltiplos xlsx, arquivos com esses tokens são ignorados
# se existir pelo menos um arquivo sem nenhum deles (o oficial).
# Adicionar novos nomes conforme aparecerem no Drive.
PERSONAL_NAME_TOKENS = {"RAFA", "GESSIANE"}

# Drive folder name → canonical client_id
# Mapeamento auditado em 2026-06-14 cruzando pastas do Drive root com core.dim_client.
# TEC PAR aponta para tecpar_edfcc744; Amigo (amigo_db1c2f0c) é sub-cliente e vive
# dentro da pasta TEC PAR — não tem pasta própria no Drive root.
# DÚVIDA ABERTA: LABTOLAB PARDINI — pode ser pardini_60395024, lab2lab_efb1cb34 ou ambos.
# PENDENTE: PHISALIA — cliente ainda não cadastrado em dim_client.
CLIENT_MAP = {
    "7K":             "bet7k_b777ab9c",
    "APERAM":         "aperam_14d1f27e",
    "CATÁLISE":       "catalise_0b7d18d6",
    "CORA":           "banco_cora_fe13d78a",
    "DAXX":           "dax_agency_00000001",
    "DOOING":         "dooing_994db77e",
    "EINSTEIN":       "einstein_6b33a588",
    "LUCKBET":        "luckbet_bea15ebc",
    "MOPAR":          "mopar_a47949f4",
    "MRV":            "mrv_f19a2136",
    "OCUPACIONAL":    "ocupacional_98c851f5",
    "PATIO MEDEIROS": "patio_medeiros_874a0358",
    "STOCCO":         "stocco_b712c66e",
    "TEC PAR":        "tecpar_edfcc744",
}

# Strategy name keywords → platform (checked via substring, case-insensitive, no accents)
PLATFORM_RULES = [
    ("retargeting", "mediasmart"),
    ("video ads", "mediasmart"),   # TODO: confirm with commercial
    ("midia programatica", "mediasmart"),
    ("midia programática", "mediasmart"),
    ("display", "mediasmart"),
    ("native ads", "mgid"),
    ("push mgid", "mgid"),
    ("push siprocal", "siprocal"),
    ("push apptargeting", "unknown"),  # TODO: confirm with commercial
    ("push", "unknown"),
]

MONTH_PT = {
    "JAN": 1, "JANEIRO": 1,
    "FEV": 2, "FEVEREIRO": 2,
    "MAR": 3, "MARCO": 3,
    "ABR": 4, "ABRIL": 4,
    "MAI": 5, "MAIO": 5,
    "JUN": 6, "JUNHO": 6,
    "JUL": 7, "JULHO": 7,
    "AGO": 8, "AGOSTO": 8,
    "SET": 9, "SETEMBRO": 9,
    "OUT": 10, "OUTUBRO": 10,
    "NOV": 11, "NOVEMBRO": 11,
    "DEZ": 12, "DEZEMBRO": 12,
}

SKIP_SHEETS = {"RESUMO", "SUMMARY", "SUGESTAO", "SUGESTOES", "SUGESTÃO", "SUGESTÕES", "INDICADORES", "TEMPLATE", "COPIA", "CÓPIA", "COPY"}


# ─── TEXT HELPERS ─────────────────────────────────────────────────────────────

def _deaccent(text: str) -> str:
    result = text.lower().strip()
    for fr, to in [
        ("á", "a"), ("à", "a"), ("â", "a"), ("ã", "a"), ("ä", "a"),
        ("é", "e"), ("è", "e"), ("ê", "e"), ("ë", "e"),
        ("í", "i"), ("ì", "i"), ("î", "i"), ("ï", "i"),
        ("ó", "o"), ("ò", "o"), ("ô", "o"), ("õ", "o"), ("ö", "o"),
        ("ú", "u"), ("ù", "u"), ("û", "u"), ("ü", "u"),
        ("ç", "c"),
    ]:
        result = result.replace(fr, to)
    return result


def _month_num(token: str) -> Optional[int]:
    t = token.upper()
    if t in MONTH_PT:
        return MONTH_PT[t]
    # Strip trailing vowel suffixes to handle forms like JUNHO→JUN, MARÇO→MAR, AGOSTO→AGO
    t2 = t.rstrip("O").rstrip("HO")
    if t2 in MONTH_PT:
        return MONTH_PT[t2]
    for length in range(len(t2), 2, -1):
        if t2[:length] in MONTH_PT:
            return MONTH_PT[t2[:length]]
    return None


def _find_quarterly_tab(wb, target_month: int) -> Optional[str]:
    """Return the sheet name matching 'PLANO JAN A MAR' / 'JAN A MAR' / etc. for target_month.
    Searches for the 'MON A MON' pattern anywhere in the tab name (handles 'PLANO' prefix).
    Returns None if no quarterly-style tab exists (e.g. V3 files with day-range tabs)."""
    for name in wb.sheetnames:
        upper = _deaccent(name).upper().strip()
        m = re.search(r"\b([A-Z]{3})\s+A\s+([A-Z]{3})\b", upper)
        if m:
            s = MONTH_PT.get(m.group(1))
            e = MONTH_PT.get(m.group(2))
            if s and e and s <= target_month <= e:
                return name
    return None


def _dates_from_drive_path(drive_path: str):
    """Derive (flight_start, flight_end) from 'YYYY/MONTHNAME' drive_folder string."""
    parts = (drive_path or "").upper().split("/")
    if len(parts) < 2:
        return None, None
    year_str = parts[0].strip()
    month_key = _deaccent(parts[1]).upper().strip()
    if not year_str.isdigit():
        return None, None
    year = int(year_str)
    month = MONTH_PT.get(month_key)
    if not month:
        return None, None
    start = date(year, month, 1)
    end = date(year, month, calendar.monthrange(year, month)[1])
    return start, end


def parse_flight_label(label: str, default_year: int = 2026):
    """
    Parse 'PLANO 11 MAI A 10 JUN' or '1 MAIO A 10 JUNHO 2026'
    into (flight_start, flight_end) or (None, None).
    """
    s = label.upper()
    s = re.sub(r"[ºª°]", "", s)
    m = re.search(r"(\d{1,2})\s+([A-ZÁÀÂÃÉÊÍÓÔÕÚÇ]{3,})\s+A\s+(\d{1,2})\s+([A-ZÁÀÂÃÉÊÍÓÔÕÚÇ]{3,})", s)
    if not m:
        return None, None
    d1, mon1, d2, mon2 = m.group(1), m.group(2), m.group(3), m.group(4)
    # Check for explicit year in the label
    year_m = re.search(r"\b(202\d)\b", s)
    year = int(year_m.group(1)) if year_m else default_year
    m1 = _month_num(mon1)
    m2 = _month_num(mon2)
    if not m1 or not m2:
        return None, None
    try:
        start = date(year, m1, int(d1))
        end_year = year + 1 if m2 < m1 else year
        end = date(end_year, m2, int(d2))
        return start, end
    except ValueError:
        return None, None


def detect_platform(strategy_name: str) -> str:
    norm = _deaccent(strategy_name)
    for keyword, platform in PLATFORM_RULES:
        if _deaccent(keyword) in norm:
            return platform
    return "unknown"


def _parse_brl(value) -> Optional[float]:
    if value is None:
        return None
    # Numeric cells from openpyxl are already parsed — don't strip their decimal points.
    # Round to 4dp to fix Excel floating-point imprecision (e.g. 8570.9999999999982 → 8571)
    if isinstance(value, (int, float)):
        v = round(float(value), 4)
        return v if v != 0 else None
    # String path: handles "R$ 1.786,00" (PT-BR: dot=thousand, comma=decimal)
    s = str(value).replace("R$", "").replace("\xa0", "").replace(" ", "")
    s = s.replace(".", "").replace(",", ".")
    s = s.strip().rstrip("%")
    try:
        return float(s) if s not in ("", "-", "***", "N/A") else None
    except ValueError:
        return None


def _parse_int(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        v = int(value)
        return v if v > 0 else None
    s = str(value).replace(",", "").replace(".", "").replace(" ", "").strip()
    if s in ("", "-", "***", "N/A"):
        return None
    try:
        return int(float(s)) or None
    except ValueError:
        return None


# ─── XLSX PARSING ─────────────────────────────────────────────────────────────

# Header keyword → internal column key
HEADER_KEYWORDS = {
    "estrategia": "strategy",
    "dispositivo": "device",
    "formato": "format",
    "modelo": "buy_model",
    "compra": "buy_model",
    "impressoes mensais": "impressions_cpm",   # CPM volume column
    "estimativa": "impressions_est",            # CPC/Push volume column
    "visualizac": "video_views",
    "investimento": "monthly_spend",
    "liquido": "monthly_spend",
    "valor": "unit_price",
    "ctr": "ctr",
    "cliques mensais": "clicks",
    "cliques": "clicks",
    "alcance": "reach",
}


def _map_headers(row) -> dict:
    col_map = {}
    for cell in row:
        if not cell.value:
            continue
        norm = _deaccent(str(cell.value))
        for kw, key in HEADER_KEYWORDS.items():
            if kw in norm and key not in col_map:
                col_map[key] = cell.column
    return col_map


def _gcell(row, col_map: dict, key: str):
    col = col_map.get(key)
    if col is None:
        return None
    return row[col - 1].value


def parse_xlsx(file_bytes: bytes, client_id: str, source_file: str,
               drive_path: str, snapshot_ts: datetime,
               default_year: int = 2026) -> list[dict]:
    """
    Parse an IO plan .xlsx.
    Returns strategy-level rows for raw.io_plan_drive_snapshot.
    default_year is extracted from the Drive folder path (e.g. 2026/MAIO → 2026).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows_out = []

    # Determine target month from drive_path (e.g. "2026/JANEIRO" → month=1)
    path_parts = (drive_path or "").upper().split("/")
    target_month = MONTH_PT.get(_deaccent(path_parts[1]).upper()) if len(path_parts) > 1 else None

    # Try to find a quarterly tab ("JAN A MAR", "ABR A JUN", etc.) for this month.
    # If found, restrict processing to that single tab and derive dates from drive_path.
    # If not found (e.g. V3 files with day-range tabs), fall through to existing logic.
    quarterly_tab = _find_quarterly_tab(wb, target_month) if target_month else None
    quarterly_dates = _dates_from_drive_path(drive_path) if quarterly_tab else (None, None)
    sheets_to_process = [quarterly_tab] if quarterly_tab else wb.sheetnames

    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        sn_upper = _deaccent(sheet_name).upper()
        if any(skip in sn_upper for skip in SKIP_SHEETS):
            continue

        # For quarterly tabs: use drive_path dates and the tab name as label.
        # For day-range tabs: extract flight label from sheet name or cell A1 (existing logic).
        if quarterly_tab:
            flight_label = sheet_name
            flight_start, flight_end = quarterly_dates
        else:
            flight_label = None
            cell_a1 = ws.cell(1, 1).value
            label_candidates = [sheet_name, str(cell_a1) if cell_a1 else ""]
            for candidate in label_candidates:
                if re.search(r"\d{1,2}\s+[A-Z]{3}", candidate.upper()):
                    flight_label = candidate
                    break

            flight_start, flight_end = None, None
            if flight_label:
                flight_start, flight_end = parse_flight_label(flight_label, default_year=default_year)

            # Drive folder = official plan for that month.
            # For day-range tabs (e.g. "11 MAI A 10 JUN"), only keep the tab if its
            # flight_start falls in the same month as drive_folder. This prevents
            # multi-month files (V3 MAI-JUL) from contributing data to months that
            # don't yet have an official plan folder in Drive.
            if target_month and flight_start and flight_start.month != target_month:
                continue

            # Final fallback: if no dates parsed from tab name, derive from drive_folder.
            # Handles files with single generic tabs (e.g. TecPar "PLANO CUIABÁ", "ABRIL", "MAIO").
            # Rule: file in a month's Drive folder = official plan for that month, regardless of tab name.
            if flight_start is None and target_month:
                flight_start, flight_end = _dates_from_drive_path(drive_path)

        # Find header row: first row where we can map ≥2 keys
        header_row = None
        col_map = {}
        for row in ws.iter_rows(min_row=1, max_row=25):
            candidate_map = _map_headers(row)
            if len(candidate_map) >= 2:
                header_row = row[0].row
                col_map = candidate_map
                break

        if not header_row or not col_map:
            continue

        # spend_type: detectado pelo unit_price das linhas CPM
        # (arquivos com nome de pessoa são filtrados antes de chegar aqui)
        spend_type_hint = None

        unit_prices_cpm = []
        strategy_rows_raw = []
        for row in ws.iter_rows(min_row=header_row + 1):
            row_text = " ".join(str(c.value) if c.value else "" for c in row).upper()
            if not any(c.value for c in row):
                continue
            if "TOTAL" in row_text:
                break
            strat_val = _gcell(row, col_map, "strategy")
            if not strat_val or str(strat_val).strip() in ("", "ESTRATÉGIAS", "ESTRATEGIAS", "OBJETIVO"):
                continue
            strategy_rows_raw.append(row)
            # Use only CPM lines for spend_type detection (CPC rates inflate the max)
            imp_cpm = _gcell(row, col_map, "impressions_cpm")
            up = _parse_brl(_gcell(row, col_map, "unit_price"))
            if imp_cpm is not None and up and up > 0:
                unit_prices_cpm.append(up)

        if spend_type_hint is not None:
            spend_type = spend_type_hint
        elif unit_prices_cpm:
            spend_type = "gross" if max(unit_prices_cpm) > 5.0 else "net"
        else:
            spend_type = "gross"

        # Build output rows
        for row in strategy_rows_raw:
            strat_name = str(_gcell(row, col_map, "strategy")).strip()
            rows_out.append({
                "client_id": client_id,
                "drive_folder": drive_path,
                "source_file": source_file,
                "spend_type": spend_type,
                "snapshot_at": snapshot_ts.isoformat(),
                "flight_label": flight_label,
                "flight_start": flight_start.isoformat() if flight_start else None,
                "flight_end": flight_end.isoformat() if flight_end else None,
                "strategy_name": strat_name,
                "platform": detect_platform(strat_name),
                "device": str(_gcell(row, col_map, "device") or "").strip() or None,
                "format": str(_gcell(row, col_map, "format") or "").strip() or None,
                "buy_model": str(_gcell(row, col_map, "buy_model") or "").strip() or None,
                "unit_price": _parse_brl(_gcell(row, col_map, "unit_price")),
                "impressions_cpm": _parse_int(_gcell(row, col_map, "impressions_cpm")),
                "impressions_est": _parse_int(_gcell(row, col_map, "impressions_est")),
                "video_views": _parse_int(_gcell(row, col_map, "video_views")),
                "monthly_spend": _parse_brl(_gcell(row, col_map, "monthly_spend")),
                "ctr_approx": _parse_brl(_gcell(row, col_map, "ctr")),
                "clicks": _parse_int(_gcell(row, col_map, "clicks")),
                "reach": _parse_int(_gcell(row, col_map, "reach")),
            })

    return rows_out


# ─── BQ OPERATIONS ────────────────────────────────────────────────────────────

def get_last_sync(bq: bigquery.Client, source_file: str, drive_folder: str) -> Optional[datetime]:
    result = list(bq.query(
        "SELECT MAX(snapshot_at) AS ts"
        " FROM `adframework.raw.io_plan_drive_snapshot`"
        " WHERE source_file = @f AND drive_folder = @d",
        job_config=bigquery.QueryJobConfig(
            query_parameters=[
                bigquery.ScalarQueryParameter("f", "STRING", source_file),
                bigquery.ScalarQueryParameter("d", "STRING", drive_folder),
            ]
        ),
    ).result())
    if result and result[0].ts:
        return result[0].ts
    return None


def insert_raw_rows(bq: bigquery.Client, rows: list[dict]):
    if not rows:
        return
    errors = bq.insert_rows_json(f"{BQ_PROJECT}.raw.io_plan_drive_snapshot", rows)
    if errors:
        raise RuntimeError(f"BQ streaming insert errors: {errors}")
    log.info(f"    → {len(rows)} rows written to raw.io_plan_drive_snapshot")


def rebuild_core_for_client(bq: bigquery.Client, client_id: str):
    """
    Full rebuild of core.io_plan_manual for a client from the latest raw snapshots.
    Aggregates all source files, separating gross vs net spend.
    Impressions: prefer net file (media platform rates); fallback to gross file.
    """
    log.info(f"    Rebuilding core.io_plan_manual for {client_id}...")

    bq.query(
        "DELETE FROM `adframework.core.io_plan_manual` WHERE client_id = @c AND plan_version = 'DRIVE-SYNC'",
        job_config=bigquery.QueryJobConfig(
            query_parameters=[bigquery.ScalarQueryParameter("c", "STRING", client_id)]
        ),
    ).result()

    insert_sql = """
        INSERT INTO `adframework.core.io_plan_manual`
            (client_id, flight_start, flight_end,
             planned_impressions, planned_clicks,
             planned_spend_gross, planned_spend_net,
             plan_version, source_file, loaded_at)

        WITH latest_snap AS (
            -- Latest snapshot_at per source_file (de-duplicate repeated syncs)
            SELECT source_file, MAX(snapshot_at) AS snap
            FROM `adframework.raw.io_plan_drive_snapshot`
            WHERE client_id = @c
              AND strategy_name NOT IN ('TOTAL', 'total')
              AND flight_start IS NOT NULL
              AND flight_end IS NOT NULL
            GROUP BY source_file
        ),
        latest AS (
            -- All strategy rows from the latest snapshot of each source_file,
            -- with DISTINCT to avoid duplicates when same file exists in multiple Drive folders
            SELECT DISTINCT r.client_id, r.source_file, r.spend_type,
                r.flight_start, r.flight_end,
                r.strategy_name, r.impressions_cpm, r.impressions_est,
                r.video_views, r.monthly_spend, r.clicks
            FROM `adframework.raw.io_plan_drive_snapshot` r
            JOIN latest_snap ls ON r.source_file = ls.source_file
                AND r.snapshot_at = ls.snap
            WHERE r.client_id = @c
              AND r.strategy_name NOT IN ('TOTAL', 'total')
              AND r.flight_start IS NOT NULL
              AND r.flight_end IS NOT NULL
        ),
        by_flight_type AS (
            SELECT
                client_id, flight_start, flight_end, spend_type,
                SUM(COALESCE(impressions_cpm, 0) + COALESCE(impressions_est, 0)) AS total_imp,
                SUM(COALESCE(clicks, 0))          AS total_clicks,
                SUM(COALESCE(monthly_spend, 0))   AS total_spend,
                STRING_AGG(DISTINCT source_file ORDER BY source_file) AS sources
            FROM latest
            GROUP BY client_id, flight_start, flight_end, spend_type
        )
        SELECT
            client_id,
            flight_start,
            flight_end,
            -- Impressions: use net file if available (media-rate file = more granular),
            -- else fall back to gross file
            COALESCE(
                MAX(CASE WHEN spend_type = 'net'   AND total_imp > 0 THEN total_imp END),
                MAX(CASE WHEN spend_type = 'gross' AND total_imp > 0 THEN total_imp END)
            )   AS planned_impressions,
            COALESCE(
                MAX(CASE WHEN spend_type = 'net'   AND total_clicks > 0 THEN total_clicks END),
                MAX(CASE WHEN spend_type = 'gross' AND total_clicks > 0 THEN total_clicks END)
            )   AS planned_clicks,
            MAX(CASE WHEN spend_type = 'gross' THEN total_spend END) AS planned_spend_gross,
            MAX(CASE WHEN spend_type = 'net'   THEN total_spend END) AS planned_spend_net,
            'DRIVE-SYNC'  AS plan_version,
            STRING_AGG(DISTINCT sources ORDER BY sources) AS source_file,
            CURRENT_TIMESTAMP() AS loaded_at
        FROM by_flight_type
        GROUP BY client_id, flight_start, flight_end
    """
    bq.query(
        insert_sql,
        job_config=bigquery.QueryJobConfig(
            query_parameters=[bigquery.ScalarQueryParameter("c", "STRING", client_id)]
        ),
    ).result()
    log.info(f"    → core.io_plan_manual rebuilt for {client_id}")


# ─── DRIVE TRAVERSAL ──────────────────────────────────────────────────────────

def _list_folder(drive_svc, folder_id: str, mime: str = None) -> list[dict]:
    query = f"'{folder_id}' in parents and trashed = false"
    if mime:
        query += f" and mimeType = '{mime}'"
    items, page_token = [], None
    while True:
        resp = drive_svc.files().list(
            q=query,
            fields="nextPageToken, files(id, name, mimeType, modifiedTime)",
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            corpora="allDrives",
        ).execute()
        items.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return items


def _has_personal_name(filename: str) -> bool:
    """True se o nome do arquivo contém token de pessoa (ex: RAFA, GESSIANE)."""
    upper = filename.upper()
    return any(token in upper for token in PERSONAL_NAME_TOKENS)


def find_plano_files(drive_svc, client_folder_id: str) -> list[dict]:
    """
    Traverse CLIENT/YEAR/MONTH/PLANO/ and return .xlsx files to sync.
    Regra de seleção por pasta PLANO:
      1. Preferir arquivos sem nome de pessoa (oficiais).
      2. Se só houver arquivos com nome de pessoa, usar como fallback.
      3. Em ambos os casos, pegar apenas o mais recente (por modifiedTime).
         Isso evita duplicação quando há múltiplas versões ou cópias renomeadas.
    Returns list of {file_id, file_name, modified_time, drive_path}.
    """
    results = []
    seen_file_ids = set()  # deduplicate: same .xlsx can appear in multiple month folders
    FOLDER_MIME = "application/vnd.google-apps.folder"
    SKIP_FOLDERS = {"DASHBOARD", "DASH", "RELATORIOS", "RELATORIO", "REPORTS", "REPORT", "OLD", "ANTIGOS"}

    for year_f in _list_folder(drive_svc, client_folder_id, FOLDER_MIME):
        if not re.fullmatch(r"\d{4}", year_f["name"].strip()):
            continue
        for month_f in _list_folder(drive_svc, year_f["id"], FOLDER_MIME):
            if month_f["name"].upper() in SKIP_FOLDERS:
                continue
            # Find PLANO subfolder (name contains "PLANO")
            for plano_f in _list_folder(drive_svc, month_f["id"], FOLDER_MIME):
                if "PLANO" not in plano_f["name"].upper():
                    continue
                xlsx_files = [
                    f for f in _list_folder(drive_svc, plano_f["id"], XLSX_MIME)
                    if f["id"] not in seen_file_ids
                ]
                if not xlsx_files:
                    continue

                official = [f for f in xlsx_files if not _has_personal_name(f["name"])]
                candidates = official if official else xlsx_files

                # Always pick only the most recent to avoid duplicating plans
                xlsx_f = max(candidates, key=lambda f: f["modifiedTime"])
                seen_file_ids.add(xlsx_f["id"])
                results.append({
                    "file_id": xlsx_f["id"],
                    "file_name": xlsx_f["name"],
                    "modified_time": xlsx_f["modifiedTime"],
                    "drive_path": f"{year_f['name']}/{month_f['name']}",
                    "drive_year": int(year_f["name"]),
                })
    return results


def download_file(drive_svc, file_id: str) -> bytes:
    buf = io.BytesIO()
    request = drive_svc.files().get_media(fileId=file_id, supportsAllDrives=True)
    dl = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return buf.getvalue()


# ─── SYNC ORCHESTRATION ───────────────────────────────────────────────────────

def sync_client(drive_svc, bq_client: bigquery.Client,
                client_name: str, folder_id: str, force: bool) -> dict:
    client_id = CLIENT_MAP[client_name]
    log.info(f"  [{client_name}] scanning Drive...")
    files = find_plano_files(drive_svc, folder_id)
    log.info(f"  [{client_name}] found {len(files)} .xlsx files across all PLANO folders")

    processed, skipped = 0, 0
    any_new = False

    for f in files:
        drive_mod = datetime.fromisoformat(f["modified_time"].replace("Z", "+00:00"))
        last_sync = get_last_sync(bq_client, f["file_name"], f["drive_path"])

        if not force and last_sync and last_sync >= drive_mod:
            log.info(f"    SKIP {f['file_name']} (last synced {last_sync.date()}, Drive mod {drive_mod.date()})")
            skipped += 1
            continue

        log.info(f"    SYNC {f['file_name']} (Drive modified {drive_mod.date()})...")
        snapshot_ts = datetime.now(timezone.utc)

        try:
            file_bytes = download_file(drive_svc, f["file_id"])
            rows = parse_xlsx(file_bytes, client_id, f["file_name"],
                              f["drive_path"], snapshot_ts,
                              default_year=f.get("drive_year", 2026))
        except Exception as e:
            log.warning(f"    ERROR parsing {f['file_name']}: {e}")
            continue

        if not rows:
            log.warning(f"    No strategy rows parsed from {f['file_name']} — skipping")
            skipped += 1
            continue

        log.info(f"    Parsed {len(rows)} strategy rows")
        insert_raw_rows(bq_client, rows)
        processed += 1
        any_new = True

    if any_new:
        rebuild_core_for_client(bq_client, client_id)

    return {"client": client_name, "processed": processed, "skipped": skipped}


# ─── MAIN ─────────────────────────────────────────────────────────────────────

OAUTH_CLIENT_FILE = r"C:\Temp\io-plan-oauth-client.json"
OAUTH_TOKEN_FILE  = r"C:\Temp\io-plan-token.json"
SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/cloud-platform",
]


def _get_credentials():
    """
    Auth strategy (in order):
    1. GOOGLE_APPLICATION_CREDENTIALS env var → service account key (local SA or CI)
    2. OAuth Desktop client file present     → OAuth installed-app flow (personal Drive)
    3. Workload identity / ADC               → Cloud Run service account
    """
    # Strategy 1: explicit SA key
    if os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"):
        creds, _ = google_auth_default(scopes=SCOPES)
        return creds

    # Strategy 2: OAuth Desktop flow (personal Drive access)
    if os.path.exists(OAUTH_CLIENT_FILE):
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request

        creds = None
        if os.path.exists(OAUTH_TOKEN_FILE):
            creds = Credentials.from_authorized_user_file(OAUTH_TOKEN_FILE, SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(OAUTH_CLIENT_FILE, SCOPES)
                creds = flow.run_local_server(port=0)
            with open(OAUTH_TOKEN_FILE, "w") as f:
                f.write(creds.to_json())
        return creds

    # Strategy 3: ADC / workload identity (Cloud Run)
    creds, _ = google_auth_default(scopes=SCOPES)
    return creds


def run_sync(client_filter: Optional[str] = None, force: bool = False) -> list[dict]:
    credentials = _get_credentials()
    drive_svc = build("drive", "v3", credentials=credentials)
    bq_client = bigquery.Client(project=BQ_PROJECT, credentials=credentials)

    all_drive_folders = _list_folder(
        drive_svc, DRIVE_ROOT_FOLDER,
        "application/vnd.google-apps.folder"
    )
    folder_by_name = {f["name"]: f["id"] for f in all_drive_folders}

    clients_to_sync = list(CLIENT_MAP.keys())
    if client_filter:
        needle = client_filter.upper()
        clients_to_sync = [c for c in CLIENT_MAP if needle in c.upper()]
        if not clients_to_sync:
            raise ValueError(
                f"No client matching '{client_filter}'. Available: {list(CLIENT_MAP.keys())}"
            )

    results = []
    for client_name in clients_to_sync:
        if client_name not in folder_by_name:
            log.warning(f"Drive folder '{client_name}' not found — skipping")
            results.append({"client": client_name, "processed": 0, "skipped": 0, "error": "folder not found"})
            continue
        result = sync_client(drive_svc, bq_client,
                             client_name, folder_by_name[client_name], force)
        results.append(result)

    return results


def main():
    parser = argparse.ArgumentParser(description="IO Plan Drive → BigQuery sync")
    parser.add_argument("--client", help="Sync only this client (e.g. cora, tecpar)")
    parser.add_argument("--force", action="store_true",
                        help="Re-sync even if Drive file unchanged")
    args = parser.parse_args()

    try:
        results = run_sync(client_filter=args.client, force=args.force)
    except ValueError as e:
        log.error(str(e))
        sys.exit(1)

    total_processed = sum(r.get("processed", 0) for r in results)
    total_skipped = sum(r.get("skipped", 0) for r in results)
    log.info(f"Done — {total_processed} files synced, {total_skipped} skipped (unchanged).")
    for r in results:
        log.info(f"  {r}")


if __name__ == "__main__":
    main()
